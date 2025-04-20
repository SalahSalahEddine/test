#flet build apk main.py
import flet as ft
import psutil
import socket
import hashlib
import based58
import platform
import datetime
import pyperclip
from flet import icons,FontWeight
from core.gorgumu._trx import *
from core.gorgumu._ransm import Encrypt
from core.gorgumu._loan import *
from core.gorgumu._client import Wallet
from core._finance import *
from core._ficher import *
from core._connexion import is_online
from PIL import Image , ImageTk
from core._rate import Gorgumu_Price
from typing import Tuple
from Crypto.Hash import RIPEMD160
from openpyxl import Workbook
from cassandra.cluster import Cluster
from decimal import Decimal,ROUND_HALF_UP
import datetime
import time
# main.py

'''
Function Include The Client Public Key And Trans It To Wallet Address With Some Algorithms Like:
    [1].Sha256
    [2].RIPEMD160
    [3].BASED58
'''
def Myaddress():
    with open("core/gorgumu/account/efa1f375d76194fa51a3556a97e641e61685f914d446979da50a551a4333ffd7.pem", "rb") as pub:
        public_key_data = pub.read()
        hashing = hashlib.sha256(str(public_key_data).encode()).hexdigest()
        hashing_to_bytes = hashing.encode()
        Ripmd = RIPEMD160.new(hashing_to_bytes)
        last_hash = Ripmd.hexdigest()
        Ripmd_to_bytes = last_hash.encode()
        enc = based58.b58encode(Ripmd_to_bytes)
        dec = enc.decode()
        return dec
'''
Function Check If Loan Is Expired
'''
def loan_is_expire():
    maddress = Myaddress()
    today = datetime.date.today()
    expire_day = session.prepare(f"SELECT expire_date FROM gorgumu.loans WHERE address = '{maddress}';")
    return expire_day == today
'''
Function To Paid Your Loan If Expire Date Is Today
    * If Client Paid,His Iscore Rise Up(This Is Helps Client For Demand Another Future Loan)
    * If Not, His Iscore Less Down (This Is Threats Client His Records)
    * The Iscore Rise & Less As Iscore Points (You Can Back To '_finance.py' File To See)
'''
def paid_bill():
    maddress = Myaddress()
    client_score = my_iscore()
    GET_CLIENT_BALANCE_QUERY = SimpleStatement(f"SELECT * FROM gorgumu.wallet WHERE address = '{maddress}';")
    rows1 = session.execute(GET_CLIENT_BALANCE_QUERY)
    for b in rows1:
        CLIENT_BALANCE = b.balance
    GET_LOAN_BOX_AMOUNT_QUERY = SimpleStatement(f"SELECT * FROM gorgumu.loanbox WHERE total = 'MAX';")
    rows2 = session.execute(GET_LOAN_BOX_AMOUNT_QUERY)
    for ls in rows2:
        LOANBOX_AMOUNT = ls.loansupply
    GET_CLIENT_DEBT_QUERY = SimpleStatement(f"SELECT * FROM gorgumu.loans WHERE address = '{maddress}';")
    rows3 = session.execute(GET_CLIENT_DEBT_QUERY)
    loan_actually_expired = loan_is_expire()
    for dbt in rows3:
        CLIENT_DEBT = dbt.debt
        #if self.loan_is_expire and CLIENT_DEBT is not None:
        if loan_actually_expired and CLIENT_DEBT is not None:
            #------------------------------|UPDATE Client Iscore|---------------------------------
            NEW_CLIENT_ISCORE = Decimal(client_score) + 10
            UPDATED_CLIENT_ISCORE = session.prepare(f"UPDATE gorgumu.wallet SET iscore = {NEW_CLIENT_ISCORE} WHERE address = '{maddress}'")
            session.execute(UPDATED_CLIENT_ISCORE)
            #------------------------------|MINUS Client Balance From Loan|-------------------------
            NEW_CLIENT_BALANCE_AFTER_LOAN_EXPIRED = Decimal(CLIENT_BALANCE) - Decimal(CLIENT_DEBT)
            CLIENT_UPDATE_BALANCE = session.prepare(f"UPDATE gorgumu.wallet SET balance = {NEW_CLIENT_BALANCE_AFTER_LOAN_EXPIRED} WHERE address = '{maddress}'")
            session.execute(CLIENT_UPDATE_BALANCE)
            #------------------------------|UPDATE MONEYSUPPLY:ADD DEBT TO MONEYSUPPLY|---------------
            NEW_MONEY_SUPPLY_AMOUNT = Decimal(LOANBOX_AMOUNT) + Decimal(CLIENT_DEBT)
            UPDATED_MONEY_SUPPLY = session.prepare(f"UPDATE gorgumu.loanbox SET loansupply = {NEW_MONEY_SUPPLY_AMOUNT} WHERE total = 'MAX'")
            session.execute(UPDATED_MONEY_SUPPLY)
            #------------------------------|UPDATE CLIENT DEBT TO NULL|--------------------------------
            UPDATED_CLIENT_DEBT = session.prepare(f"UPDATE gorgumu.loans SET debt = 0 WHERE address = '{maddress}'")
            session.execute(UPDATED_CLIENT_DEBT)
            print('Expired,Your Loan Date Expire Is Today,We Get Back The Amount Succufuly')
        elif CLIENT_BALANCE is None or CLIENT_BALANCE < CLIENT_DEBT and client_score is not None:
            #------------------------------|UPDATE CLIENT ISCORE|---------------------------------------
            NEW_CLIENT_ISCORE = Decimal(client_score) - 50
            UPDATED_CLIENT_ISCORE = session.prepare(f"UPDATE gorgumu.wallet SET iscore = {NEW_CLIENT_ISCORE} WHERE address = '{maddress}'")
            session.execute(UPDATED_CLIENT_ISCORE)
            #-------------------|Start Encryption|-----------
            Encrypt()
        else:
            pass
'''
Function Check The Client Wallet Expire Date Is Today
'''
def balance_is_expire():
    maddress = Myaddress()
    today = datetime.date.today()
    expire_day = session.prepare(f"SELECT expire_date FROM gorgumu.wallet WHERE address = '{maddress}';")
    return expire_day == today
'''
Function Update The Client Balance To Zero If The Expire Date Is Today
'''
def null_balance():
    maddress = Myaddress()
    #if self.balance_is_expire():
    expire = balance_is_expire()
    if expire == True:
        update_query = session.prepare(f"UPDATE gorgumu.wallet SET balance = 0 WHERE address = '{maddress}';")
        session.execute(update_query)
        print('Balance,Your Balance Null Because Your Expire Date Is Today!')
        print("Your Balance Is Expired")
    else:
        print("The Expire Date Not Comming Yet!")

'''
Function To Create Excel File And Store The GRM Price Every Single Day
'''
def update_records():
        filename = "Records.xlsx"
        filename_path = "core/gorgumu/"
        file_path = os.path.join(filename_path, filename)
        last_update_file = "last_update.txt"
        last_update_file_path = "core/gorgumu/"
        last_update_path = os.path.join(last_update_file_path, last_update_file)
        current_date = datetime.datetime.now().date()
        if os.path.exists(last_update_path):
            with open(last_update_path, 'r') as f:
                last_update = datetime.datetime.strptime(f.read(), '%Y-%m-%d').date()
        else:
            with open(last_update_path, 'w') as f:
                f.write(str(current_date))
            last_update = None
        if current_date != last_update:
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
            else:
                wb = Workbook()
            ws = wb.active
            if last_update is None:
                ws["A1"] = "DATE"
                ws["B1"] = "PRICE"
            ws.append([str(current_date), Gorgumu_Price])
            wb.save(file_path)
            with open(last_update_path, 'w') as f:
                f.write(str(current_date))
def main(gorgumu:ft.Page):
    gorgumu.title='Gorgumu'
    gorgumu.theme_mode = ft.ThemeMode.LIGHT
    gorgumu.window.width = 390
    gorgumu.window.height = 740
    gorgumu.window.resizable=False
    gorgumu.window.title_bar_hidden = True
    gorgumu.horizontal_alignment = 'center'
    gorgumu.vertical_alignment = 'center'
    gorgumu.scroll = 'auto'
    def root_change(route):
        d = datetime.datetime.now()
        global signinpasswd
        signinpasswd = ft.TextField(label='Password:',password=True,can_reveal_password=True)
        gorgumu.views.clear()
        gorgumu.views.append(
            ft.View('/',
                    [
                       ft.Text('CBDC System',color='black'),
                       signinpasswd,
                       ft.TextButton('Login',icon=icons.LOGIN_OUTLINED,on_click=login_method),
                       ft.TextButton('Create New Wallet',on_click=lambda _: gorgumu.go('/signup'),icon=icons.WALLET),
                       ft.Text('All Rights Reserved',color='whitesmoke',size=10),
                       ft.Text(d.year,color='whitesmoke',size=10)
                       

                    ],horizontal_alignment='center',vertical_alignment='center',bgcolor='whitesmoke')
        )
        
        if gorgumu.route == '/signup':
            global signuppasswd1
            global signuppasswd2
            signuppasswd1 = ft.TextField(label='Password:',password=True,can_reveal_password=True)
            signuppasswd2 = ft.TextField(label='Password:',password=True,helper_text='must be strong',can_reveal_password=True)
            gorgumu.views.append(
            ft.View('/signup',
                    [
                       signuppasswd1,
                       signuppasswd2,
                       ft.TextButton('Register',on_click=signup_method),
                       ft.TextButton('Back',on_click=page_go,icon=icons.TURN_LEFT)
                       

                    ],horizontal_alignment='center',vertical_alignment='center',bgcolor='whitesmoke')
        )
        if gorgumu.route == '/wallet':
            update_records()
            paid_bill()
            null_balance()
            GRM = grm_balance()
            rubel_balance = usd_balance()
            gorgumu.views.append(
            ft.View('/wallet',
                    [

                       #ft.Container(bgcolor='red',width=125,height=123,content=ft.Image(src='core/gorgumu/account/61f2c041d4e9e0f558ae5ee8d6adb62e79bacce45be6d0b6d0bbe8947f5f0dd9.png',width=80,height=80)),
                       #ft.PopupMenuButton(items=[ft.PopupMenuItem(text='Settings'),ft.PopupMenuItem(text='Logout'),ft.PopupMenuItem(text='About')]),
                       #ft.AppBar(bgcolor='white',actions=[ft.PopupMenuButton(items=[ft.PopupMenuItem(text='Settings'),ft.PopupMenuItem(text='Logout'),ft.PopupMenuItem(text='About')])]),
                       ft.Image(src='core/gorgumu/account/61f2c041d4e9e0f558ae5ee8d6adb62e79bacce45be6d0b6d0bbe8947f5f0dd9.png',width=80,height=80),
                       ft.Text('Total :',color='gray',size=12),
                       ft.Text(f'{rubel_balance} $',color='black',weight=FontWeight.BOLD),
                       ft.Text('Balance :',color='gray',size=12),
                       ft.Text(f'{GRM} GRM',color='black',weight=FontWeight.BOLD),
                       ft.FloatingActionButton(bgcolor='white',icon=ft.Icons.SEND,on_click=lambda _: gorgumu.go('/trx'),height=50,width=50),
                       #ft.CupertinoNavigationBar(bgcolor='white',inactive_color='gray',destinations=[ft.NavigationBarDestination(icon=icons.RECEIPT,label='Receive'),ft.NavigationBarDestination(icon=icons.ADD,label='Send'),ft.NavigationBarDestination(icon=icons.HISTORY,label='History')]),
                       ft.FloatingActionButton(bgcolor='white',icon=ft.Icons.VERIFIED_USER,on_click=lambda _: gorgumu.go('/receive'),height=50,width=50),
                       ft.FloatingActionButton(bgcolor='white',icon=ft.Icons.SETTINGS,on_click=lambda _: gorgumu.go('/settings'),height=50,width=50),
                       ft.FloatingActionButton(bgcolor='white',icon=ft.Icons.HANDSHAKE,on_click=lambda _: gorgumu.go('/loan'),height=50,width=50),
                       ft.FloatingActionButton(bgcolor='white',icon=ft.Icons.ADD,on_click=lambda _: gorgumu.go('/deposite'),height=50,width=50),
                       ft.FloatingActionButton(bgcolor='white',icon=icons.HISTORY,on_click=lambda _: gorgumu.go('/history')),
                       ft.FloatingActionButton(bgcolor='white',icon=ft.Icons.LOGOUT,on_click=signout)
                       

                    ],
                    horizontal_alignment='center',
                    bgcolor='whitesmoke')
        )
        if gorgumu.route == '/trx':
            global receiver_address
            global amount_to_send
            global currency_dropdown
            receiver_address = ft.TextField(label='Receiver:',border_color='white')
            amount_to_send = ft.TextField(label='Amount:',border_color='white')
            currency_dropdown = ft.DropdownM2(options=[ft.dropdownm2.Option("GRM"),ft.dropdownm2.Option("ZYN")],)
            gorgumu.views.append(
            ft.View('/trx',
                    [
                       ft.AppBar(bgcolor='white',center_title=True),
                       receiver_address,
                       amount_to_send,
                       currency_dropdown,
                       ft.TextButton('SEND',on_click=send_method),
                       ft.TextButton('Back',on_click=lambda _: gorgumu.go('/wallet'),on_hover='whitesmoke')
                    ],horizontal_alignment='center',bgcolor='whitesmoke')
        )
        if gorgumu.route == '/receive':
            global wallet_address
            wallet_address = Myaddress()
            prv_key = Private_key()
            gorgumu.views.append(
            ft.View('/receive',
                    [
                       ft.Image(src='core/gorgumu/account/61f2c041d4e9e0f558ae5ee8d6adb62e79bacce45be6d0b6d0bbe8947f5f0dd9.png',width=180,height=180),
                       ft.AppBar(bgcolor='white',center_title=True),
                       ft.TextField(label='Wallet Address:',border_color='white',read_only=True,value=wallet_address,on_click=copyaddress),
                       ft.TextField(label='Private Key:',border_color='white',read_only=True,bgcolor='gray',helper_text='Dont Share This Key',value=prv_key),
                       ft.TextButton('Back',on_click=lambda _: gorgumu.go('/wallet'),on_hover='whitesmoke')
                    ],horizontal_alignment='center',bgcolor='whitesmoke')
        )
        if gorgumu.route == '/history':
            gorgumu.views.append(
            ft.View('/history',
                    [
                       ft.DataTable(
                                    columns=[
                                        ft.DataColumn(label=ft.Text("Sender")),
                                        ft.DataColumn(label=ft.Text("Receiver")),
                                        ft.DataColumn(label=ft.Text("Amount"))],
                                    rows=[
                                        ft.DataRow( cells=[
                                                            ft.DataCell(ft.Text("Salah")),
                                                            ft.DataCell(ft.Text("Oussama")),
                                                            ft.DataCell(ft.Text("666")),
                                                    ])
                                    ]),
                        ft.TextButton('Back',on_click=lambda _: gorgumu.go('/wallet'),on_hover='whitesmoke')
                    ],horizontal_alignment='center',bgcolor='whitesmoke')
            )
        if gorgumu.route == '/settings':
            global newpasswd1
            global newpasswd2
            newpasswd1 = ft.TextField(label='New Password:',border_color='white',password=True,can_reveal_password=True)
            newpasswd2 = ft.TextField(label='Repeat New Password:',border_color='white',password=True,can_reveal_password=True)
            gorgumu.views.append(
            ft.View('/settings',
                    [
                       ft.AppBar(bgcolor='white',center_title=True),
                       newpasswd1,
                       newpasswd2,
                       ft.TextButton('Submit',on_click=change_password),
                       ft.TextButton(icon=icons.DELETE,on_click=delete_wallet),
                       ft.TextButton('Back',on_click=lambda _: gorgumu.go('/wallet'),on_hover='whitesmoke')

                    ],horizontal_alignment='center',bgcolor='whitesmoke')
        
        )
        if gorgumu.route == '/loan':
            global Loanamount
            Loanamount = ft.TextField(label='Amount:',border_color='white')
            gorgumu.views.append(
            ft.View('/loan',
                    [
                       Loanamount,
                       ft.Text('1-At Least 20 Transactions',color='red'),
                       ft.Text('2-Max Day Limit : 60 Days',color='red'),
                       ft.Text('3-Type Loan:Ninja (Client Data is Asset)',color='red'),
                       ft.Text('4-Fractional Reserve:10%',color='red'),
                       ft.TextButton('Demand',on_click=ask_loan),
                       ft.TextButton('Back',on_click=lambda _: gorgumu.go('/wallet'),on_hover='whitesmoke')

                    ],horizontal_alignment='center',bgcolor='whitesmoke')
        
        )
        if gorgumu.route == '/deposite':
            global depositeamount
            depositeamount = ft.TextField(label='Amount:',border_color='white')
            gorgumu.views.append(
            ft.View('/deposite',
                    [
                       ft.AppBar(bgcolor='white',center_title=True),
                       depositeamount,
                       ft.TextButton('Deposit',on_click=deposite_method),
                       ft.TextButton(icon=icons.MONEY,on_click=get_back_cash),
                       ft.TextButton('Back',on_click=lambda _: gorgumu.go('/wallet'),on_hover='whitesmoke')

                    ],horizontal_alignment='center',bgcolor='whitesmoke')
        
        )
        
        gorgumu.update()
    '''
    Function To Create a New Wallet
    '''
    def signup_method(e):
        Sign = Wallet()
        x = signuppasswd1.value
        if not signuppasswd1.value and not signuppasswd2.value:
            return False
        if signuppasswd1.value == signuppasswd2.value:
            Sign.save_keys()
            Sign.encrypt_password(x)
            Sign.register_wallet()
            time.sleep(5)
            Sign.myqrcode()
            gorgumu.go('/wallet')
    '''
    Function To Log The Client In Wallet After Type His Own Password
    '''
    def login_method(e):
        psl = signinpasswd.value
        with open("core/gorgumu/account/e7cf3ef4f17c3999a94f2c6f612e8a888e5b1026878e4e19398b23bd38ec221a.ini",'r') as psw:
            readpasswd = psw.read()
        Mac_address = None
        Hostname = socket.gethostname()
        Operation_System = platform.system()
        Passwd = str(psl) + str(Mac_address) + str(Hostname) + str(Operation_System)
        hash = hashlib.sha256(Passwd.encode()).hexdigest()
        if readpasswd == hash:
            print("Login Succufully")
            #------------------------------------|Save Session File Again To Keep App Open When it Found|-----------------------
            today = datetime.datetime.now()
            with open("core/gorgumu/account/3f3af1ecebbd1410ab417ec0d27bbfcb5d340e177ae159b59fc8626c2dfd9175.tmp","w") as Session:
                Session.write(str(today))
            gorgumu.go('/wallet')
        else:
            print("Login Failed")
    '''
    Function Include The Client Private Key And Trans It  With Some Algorithms Like:
    [1].Sha256
    [2].RIPEMD160
    [3].BASED58
    '''
    def Private_key():
        with open("core/gorgumu/account/715dc8493c36579a5b116995100f635e3572fdf8703e708ef1a08d943b36774e.pem","rb") as prv:
            private_key_data = prv.read()
            hashingx = hashlib.sha256(str(private_key_data).encode()).hexdigest()
            hashing_to_bytes = hashingx.encode()
            Ripmd = RIPEMD160.new(hashing_to_bytes)
            last_hash = Ripmd.hexdigest()
            Ripmd_to_bytes = last_hash.encode()
            enc = based58.b58encode(Ripmd_to_bytes)
            dec = enc.decode()
            return dec
    '''
    Function To Client Can Copy His Wallet Address For Give it To Somebody For Receive Balance From Him
    '''
    def copyaddress(e):
        pyperclip.copy(wallet_address)
    '''
    Function To Calculate Client Balance On USD
    For Calculate How Much USD Just Do This: (GRM BALANCE * GRM PRICE)

    '''
    def usd_balance():
        #-----------------------|Get The Client Balance|-------------------
        address = Myaddress()
        query = SimpleStatement(f"SELECT balance FROM gorgumu.wallet WHERE address = '{address}' ;")
        rows = session.execute(query)
        for row in rows:
            balance_on_Grm = row.balance
        #----------------------|Start Calculate The Balance On USD|--------------------------
        balance_on_usd = Decimal(balance_on_Grm) * Decimal(Gorgumu_Price)
        formated_usd_balance = balance_on_usd.quantize(Decimal('1E-4'),rounding=ROUND_HALF_UP)
        return formated_usd_balance
    '''
    Function To Calculate Client Balance On GRM
    '''
    def grm_balance():
        address = Myaddress()
        query = SimpleStatement(f"SELECT balance FROM gorgumu.wallet WHERE address = '{address}' ;")
        rows = session.execute(query)
        for row in rows:
            balance_on_Grm = row.balance
        return balance_on_Grm
    '''
    Function To Let Client Changes The Wallet Password
    '''
    def change_password(e):
        npasswd = newpasswd1.value
        repasswd = newpasswd2.value
        if npasswd == repasswd:
            Mac_address = None
            Hostname = socket.gethostname()
            Operation_System = platform.system()
            Passwd = str(npasswd) + str(Mac_address) + str(Hostname) + str(Operation_System)
            hash = hashlib.sha256(Passwd.encode()).hexdigest()
            with open("core/gorgumu/account/e7cf3ef4f17c3999a94f2c6f612e8a888e5b1026878e4e19398b23bd38ec221a.ini",'w') as newlongpassword:
                newlongpassword.write(hash)
            file_tmp_path = 'core/gorgumu/account/3f3af1ecebbd1410ab417ec0d27bbfcb5d340e177ae159b59fc8626c2dfd9175.tmp'
            os.remove(file_tmp_path)
            print('password has been changed')
            gorgumu.views.clear()
            gorgumu.go('/')
    '''
    Function Submit The Transaction After Filled The Correct Address & Amount
    **Neccessary Available Internet Connection
    Notice: Before Sent Transaction You Have To Know That One GRM Unit Equal 1024 ZYN
    '''
    def send_method(e):
        connect = is_online()
        try:
            if connect:
                maddress = Myaddress()
                exchange_rate = {"GRM":1,"ZYN":1024}
                currency_selected = currency_dropdown.value
                receipent = receiver_address.value
                amountt = float(amount_to_send.value)
                if receipent == maddress:
                    print('Failed Send Transaction')
                elif currency_selected == "GRM":
                    will_send_grm = amountt * exchange_rate["GRM"]
                    tx1 = Transaction(receipent,will_send_grm)
                    tx1.calculate_hash()
                    tx1.sign_transaction(tx1.data)
                    tx1.Send(tx1.data,tx1.receiver,tx1.amount)
                    time.sleep(2)
                    gorgumu.views.clear()
                elif currency_selected == "ZYN":
                    will_send_zyn = amountt / exchange_rate["ZYN"]
                    tx2 = Transaction(receipent,will_send_zyn)
                    tx2.calculate_hash()
                    tx2.sign_transaction(tx2.data)
                    tx2.Send(tx2.data,tx2.receiver,tx2.amount)
                    time.sleep(2)
                    gorgumu.views.clear()
            else:
                print("Error Connection","Check Your Internet Connection")
        except:
            pass
    '''
    Function to calculate loansupply
    '''
    def loansupply():
        loanbox_query = SimpleStatement("SELECT * FROM gorgumu.loanbox;")
        rows = session.execute(loanbox_query)
        for rs in rows:
            loansupply = rs.loansupply
            return loansupply
    '''
    Function Give Client a Method To Get Loan Type Ninja With Conditions
    '''
    def ask_loan(e):
        try:
            maddress = Myaddress()
            ls = loansupply()
            #Allow Client To Get Just Half Of Total Loansupply
            global halfsupply
            halfsupply = ls / 2
            loanamountvalue = Decimal(Loanamount.value)
            if loanamountvalue <= halfsupply:
                return calculate_loan(maddress,loanamountvalue)
            else:
                print('Warning,Something Wrong! We Dont Have This Supply')
        except:
            pass
    '''
    Function To Delete Wallet Address
    '''
    def delete_wallet(e):
        maddress = Myaddress()
        credentials_folder_path = 'core/gorgumu/account'
        response = messagebox.askyesno("Delete Wallet","Do You Want To Remove Your Wallet Serioucely!?")
        if response:
            #----------------------|Delete Wallet From Database|--------------------------
            delete_statement = session.prepare("DELETE FROM gorgumu.wallet WHERE address = ?")
            session.execute(delete_statement, (maddress,))
        #-----------|Delete Keys And Passwords|--------------
        creds_files = os.listdir(credentials_folder_path)
        for cred in creds_files:
            cred_path = os.path.join(credentials_folder_path, cred)
            os.remove(cred_path)
        gorgumu.views.clear()
    '''
    Function Let Client Deposite His Amount For Store It
    The System Use Fractional Reserve
    '''
    def deposite_method(e):
        maddress = Myaddress()
        query = SimpleStatement(f"SELECT balance FROM gorgumu.wallet WHERE address = '{maddress}' ;")
        rows = session.execute(query)
        for row in rows:
            Obalance = row.balance 
        my_amount_to_deposit = Decimal(depositeamount.value)
        #-------------|Check To Confirm My Amount Wanna Deposit it Dosen't Big Then My Balance|---------
        if my_amount_to_deposit > Decimal(Obalance):
            print('You Dont Have This Balance!')
        #-------------|Return False When Amount Is Null|---------------------------
        elif my_amount_to_deposit == 0:
            print('Error,You Cant Enter Null Amount.Try Again')
        #--|If True Deposit , Insert Wallet Address & Amount To Deposit Table--|
        elif my_amount_to_deposit <= Obalance:
            #---------------------|UPDATE Client Balance|-----------------------
            NEW_CLIENT_BALANCE = Decimal(Obalance) - Decimal(my_amount_to_deposit)
            QUERY = SimpleStatement(f"UPDATE gorgumu.wallet SET balance = {NEW_CLIENT_BALANCE} WHERE address = '{maddress}'")
            session.execute(QUERY)
            #-------------|CALCULATE 10% FROM DEPOSITE Amount FOR Fractional Reserve Amount|------
            FRACTIONAL_RESERVE_RATE = 10
            CALCULATE_PERCENTAGE = FRACTIONAL_RESERVE_RATE / 100
            FRACTIONAL_RESERVE = Decimal(my_amount_to_deposit) * Decimal(CALCULATE_PERCENTAGE)
            #------------------|CALCULATE LOAN AMOUNT|-----------------
            LOAN_VALUE = Decimal(my_amount_to_deposit) - FRACTIONAL_RESERVE
            FRACTIONAL_RESERVE_RESULT = Decimal(my_amount_to_deposit) - LOAN_VALUE
            #------------------|SELECT LOANBOX & FRACTIONAL_RESERVE|----------------
            get_loanbox_amount = SimpleStatement("SELECT * FROM gorgumu.Loanbox WHERE total = 'MAX';")
            row = session.execute(get_loanbox_amount)
            for x in row:
                loanbox_amout = x.loansupply
            get_fractional_reserve_amount = SimpleStatement("SELECT * FROM gorgumu.Fractional_Reserve WHERE total = 'MAX';")
            rows = session.execute(get_fractional_reserve_amount)
            for ys in rows:
                fractional_reserve_x = ys.money_supply
            #-----------------------------|INSERT Deposit Amount To "DEPOSITE" TABLE|--------------------------
            INSERT_DEPOSITE = session.prepare(f"INSERT INTO gorgumu.DEPOSITE (address,amount) VALUES ('{maddress}',{my_amount_to_deposit})")
            session.execute(INSERT_DEPOSITE)
            #--------|CALCULATE 10% FROM DEPOSITE Amount FOR Fractional Reserve Amount|-------
            FRACTIONAL_RESERVE_RATE = 10
            CALCULATE_PERCENTAGE = FRACTIONAL_RESERVE_RATE / 100
            FRACTIONAL_RESERVE = Decimal(my_amount_to_deposit) * Decimal(CALCULATE_PERCENTAGE)
            #|-------|CALCULATE LOAN AMOUNT|-------------------
            LOAN_VALUE = my_amount_to_deposit - FRACTIONAL_RESERVE
            UPDATE_LOAN_VALUE = loanbox_amout + Decimal(LOAN_VALUE)
            #|-----------------------------------|INSERT FRACTIONAL RESERVE & LOAN INTO TABLES|-----------------------
            INSERT_LOAN = session.prepare(f"UPDATE gorgumu.loanbox SET loansupply = {UPDATE_LOAN_VALUE} WHERE total = 'MAX';")
            session.execute(INSERT_LOAN)
            FRACTIONAL_RESERVE_RESULT = my_amount_to_deposit - LOAN_VALUE
            UPDATED_FRACTIONAL_RESERVE = fractional_reserve_x + Decimal(FRACTIONAL_RESERVE_RESULT)
            INSERT_FRACTIONAL_RESERVE = session.prepare(f"UPDATE gorgumu.Fractional_Reserve SET money_supply = {UPDATED_FRACTIONAL_RESERVE} WHERE total = 'MAX';")
            session.execute(INSERT_FRACTIONAL_RESERVE)
            time.sleep(1)
            gorgumu.views.clear()
            print('Deposit,Deposit Succufull Thank You!')
    '''
    Function To Get Back The Cash That Client Deposite If Available On Reserve
    '''
    def get_back_cash(e):
        maddress = Myaddress()
        #--------------------------|Get Client Balance|-------------------------------
        query = SimpleStatement(f"SELECT * FROM gorgumu.wallet WHERE address = '{maddress}' ;")
        rows = session.execute(query)
        for b in rows:
            my_balance = b.balance
        #---------|Get The Amount That Client Deposite From DEPOSITE Table|-----------------
        query1 = SimpleStatement(f"SELECT amount FROM gorgumu.DEPOSITE WHERE address = '{maddress}' ;")
        rows1 = session.execute(query1)
        for r in rows1:
            MCash = r.amount
        #-------|Check How Much Amount On Fractional Reserve|--------
        query2 = SimpleStatement("SELECT * FROM gorgumu.FRACTIONAL_RESERVE ;")
        rows2 = session.execute(query2)
        for supply in rows2:
            #Whole FRACTIONAL RESERVE Amount
            Total_Reserve = supply.money_supply
        #Check If Client Deposit Amount Not None
        if MCash is not None and MCash < Total_Reserve:
            BACKCASH = Total_Reserve - Decimal(MCash)
            NEWBALANCE = Decimal(my_balance) + Decimal(MCash)
            #-----------------------------------------------|UPDATE CLIENT BALANCE|--------------------------------
            query3 = session.prepare(f"UPDATE gorgumu.wallet SET balance = {NEWBALANCE} WHERE address = '{maddress}'")
            session.execute(query3)
            #-----------------------------------------------|UPDATE FRACTIONAL_RESERVE|----------------------------
            FRACTIONAL_RESERVE_UPDATE_QUERY = session.prepare(f"UPDATE gorgumu.FRACTIONAL_RESERVE SET money_supply = {BACKCASH} WHERE total = 'MAX'")
            session.execute(FRACTIONAL_RESERVE_UPDATE_QUERY) 
            print('Succuful,Your Amount That You Deposit Has Been Whitrawed Succufuul.Thank You For Trust Us')
        if MCash == '':
            print('Failed","Your Amount Didnt Found Or Dosent Available Currently!')
    '''
    Function To Sign out From Wallet With Delete One tmp File
    '''
    def signout(e):
        file_path = "core/gorgumu/account/3f3af1ecebbd1410ab417ec0d27bbfcb5d340e177ae159b59fc8626c2dfd9175.tmp"
        try:
            os.remove(file_path)
            print(f"File '{file_path}' deleted successfully.")
            gorgumu.views.clear()
            gorgumu.go('/')
        except FileNotFoundError:
            print(f"File '{file_path}' not found.")
    def page_go(view):
        gorgumu.views.pop()
        backPage = gorgumu.views[-1]
        gorgumu.go(backPage.route)

    gorgumu.on_route_change = root_change
    gorgumu.on_view_pop = page_go
    gorgumu.go(gorgumu.route)

ft.app(main)