from .gorgumu._db import *
from dotenv import load_dotenv # type: ignore
from decimal import Decimal
from openpyxl import load_workbook # type: ignore
from ._rate import Gorgumu_Price,number_of_units

#---|Calculate Change Rate of Change|-------------
def calculate_rate_of_change():
    #-----------|Include Excel File|--------------
    wb = load_workbook("core/gorgumu/Records.xlsx")
    worksheet = wb.active
    prices = []
    for x in worksheet.iter_rows(min_row=2, min_col=2,values_only=True):
        prices.append(x[0])
    R = []
    if prices:
        for n in range(len(prices) -1):
            try:
                divisor = prices[n]
                result = (prices[n + 1] / divisor) -1
                R.append(result * 100)
                return R
            except ZeroDivisionError:
                return None
    else:
        return []

#-------|Calculate Inflation Rate|--------
def calculate_inflation():
    '''
    Function To Calculate Inflation

    Variables:

    [1].R_EXCHANGE: Variable To Call (calculate_rate_of_change()) Function
    [2].INFLATION: Variable To Sum List Value Inside The Previous Function With (R_EXCHANGE) Variable
    '''
    R_EXCHANGE = calculate_rate_of_change()
    if R_EXCHANGE:
        INFLATION = sum(R_EXCHANGE)
        return INFLATION
    else:
        return 0
#-------|Calculate The Nominal Interest|--------------------------
def calculate_nominal_interest(**args):
    π = calculate_inflation()
    if "principal" in args:
        Nominal_Interest =  args["principal"] * Decimal(2/12) * Decimal(0.05)
        return Nominal_Interest
    #---|Control The Interest Rate|--------------------------
    if π > 5:
        '''
        Rise Up Nominal Interest When Inflation More Than 5 %
        '''
        return Nominal_Interest + π
    if π < 2:
        '''
        Down Nominal Interest When Inflation Less Than 2 %
        '''
        return Nominal_Interest - π

#---|Calculate The Real Interest Rate|-----
def calculate_real_intrest(loan):
    '''
    Function To Calculate The Real Interest:
    
    Variables:

    [1].π: Variable To Call (calculate_inflation) To Return The Inflation Rate
    [2].N_Rate: Variable To Call (calculate_nominal_interest()) To Return Nominal Interest Rate

    Args:
    [1].loan = Arg Of Loan Amount

    To Calculate Rel Interest: Minus The Nominal Interest Variable From Inflation Rate Variable
    '''
    π = calculate_inflation()   # π: Inflation
    N_Rate = calculate_nominal_interest(principal=loan) #Nominal Rate
    return N_Rate - Decimal(π)

#-|Generate Liquidity|-
def generate_liquidity():
    '''
    Function To Calculate New GRM Supply If Current Circulating Supply Under 10K

    Variables:

    [1].π: Variable To Call (calculate_inflation) To Return The Inflation Rate
    [2].circulating_supply: Variable To Get How Much Supply Available
    [3].unit_price: Variable To Get GRM Current Price
    [4].unit_price_with_inflation: Varaiable To Get GRM Price With Count Inflation Rate
    [5].new_units: Variable To Store The Result Of New GRM Units To Create
    '''
    π = calculate_inflation()
    unit_price = Gorgumu_Price
    #Calculate GRM Price With Count Inflation
    unit_price_with_inflation = unit_price * (1 +  π)
    #Check How Much GRM Supply Units Remain
    circulating_supply = number_of_units
    #Calulate How Much GRM Unites Purchaced On Circulating Supply If Remains Under 10k
    if circulating_supply <= 10000:
        new_units = circulating_supply / unit_price_with_inflation * 4
        return new_units