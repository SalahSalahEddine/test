import os
import PIL
import time
import qrcode
import zmq
import ecdsa
import re
import psutil
import socket
import hashlib
import based58
import platform
import datetime
import pyperclip
import PIL.Image
import tkinter as tk
from threading import Thread
from tkinter import *
from tkinter import ttk
from time import strftime
import customtkinter as ctk
from .gorgumu._db import *
from .gorgumu._trx import *
from .gorgumu._ransm import Encrypt
from .gorgumu._loan import *
from .gorgumu._client import Wallet
from ._finance import *
from ._ficher import *
from ._connexion import is_online
from tkinter import messagebox
from PIL import Image , ImageTk
from ._rate import Gorgumu_Price
from typing import Tuple
from Crypto.Hash import RIPEMD160
from openpyxl import Workbook
from cassandra.cluster import Cluster
from decimal import Decimal,ROUND_HALF_UP
from tkinter import messagebox
'''
Function Include The Client Public Key And Trans It To Wallet Address With Some Algorithms Like:
[1].Sha256
[2].RIPEMD160
[3].BASED58
'''
def Myaddress():
    with open("core/gorgumu/account/efa1f375d76194fa51a3556a97e641e61685f914d446979da50a551a4333ffd7.pem", "rb") as pub:
        public_key_data = pub.read()
        hashing = hashlib.sha256(str(public_key_data).encode()).hexdigest()
        hashing_to_bytes = hashing.encode()
        Ripmd = RIPEMD160.new(hashing_to_bytes)
        last_hash = Ripmd.hexdigest()
        Ripmd_to_bytes = last_hash.encode()
        enc = based58.b58encode(Ripmd_to_bytes)
        dec = enc.decode()
        return dec
'''
Function Include The Client Private Key And Trans It  With Some Algorithms Like:
[1].Sha256
[2].RIPEMD160
[3].BASED58
'''
def Private_key():
    with open("core/gorgumu/account/715dc8493c36579a5b116995100f635e3572fdf8703e708ef1a08d943b36774e.pem","rb") as prv:
            private_key_data = prv.read()
            hashingx = hashlib.sha256(str(private_key_data).encode()).hexdigest()
            hashing_to_bytes = hashingx.encode()
            Ripmd = RIPEMD160.new(hashing_to_bytes)
            last_hash = Ripmd.hexdigest()
            Ripmd_to_bytes = last_hash.encode()
            enc = based58.b58encode(Ripmd_to_bytes)
            dec = enc.decode()
            return dec
#Wallet Window
class Main(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Gorgumu v 1.0")
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("dark-blue")
        self.window_width = 900
        self.window_height = 515
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.x = int((self.screen_width/2) - (self.window_width/2))
        self.y = int((self.screen_height/2) - (self.window_height/2))
        self.geometry(f"{self.window_width}x{self.window_height}+{self.x}+{self.y}")
        self.iconbitmap("core/icon/gorgumu_icon.ico")
        self.resizable(False, False)
        self.baseframe = self.mainframe()
        self.sideframe = self.sidebarframe()
        self.flag = True
        self.excel = self.update_records()
        self.bill = self.paid_bill()
        self.null_balance()
        self.mainloop()
    #----------------------
    #-------|Widgets|------
    #----------------------
    '''
    Function To Calculate Client Balance On USD
    For Calculate How Much USD Just Do This: (GRM BALANCE * GRM PRICE)

    '''
    def usd_balance(self):
        #-----------------------|Get The Client Balance|-------------------
        address = Myaddress()
        query = SimpleStatement(f"SELECT balance FROM gorgumu.wallet WHERE address = '{address}' ;")
        rows = session.execute(query)
        for row in rows:
            balance_on_Grm = row.balance
        #----------------------|Start Calculate The Balance On USD|--------------------------
        balance_on_usd = Decimal(balance_on_Grm) * Decimal(Gorgumu_Price)
        formated_usd_balance = balance_on_usd.quantize(Decimal('1E-4'),rounding=ROUND_HALF_UP)
        return formated_usd_balance
    '''
    Function To Calculate Client Balance On GRM
    '''
    def grm_balance(self):
        address = Myaddress()
        query = SimpleStatement(f"SELECT balance FROM gorgumu.wallet WHERE address = '{address}' ;")
        rows = session.execute(query)
        for row in rows:
            balance_on_Grm = row.balance
        return balance_on_Grm
    #Mainframe
    def mainframe(self):
        #------------------|Menu|------------------------------------------------
        menubar = Menu(self)
        depostit_menu = Menu(menubar,tearoff=0)
        depostit_menu.add_command(label="Deposit",command=self.call_deposite_window)
        depostit_menu.add_separator()
        menubar.add_cascade(label="Service",menu=depostit_menu)
        operation = Menu(menubar, tearoff=0)
        operation.add_command(label="Perfomence",command=self.call_perfomence_window)
        operation.add_command(label="Transactions",command=self.call_ledger_window)
        operation.add_separator()
        menubar.add_cascade(label="Options",menu=operation)
        #===================|Help Menu|---------------------------------
        helpmenu = Menu(menubar, tearoff=0)
        helpmenu.add_command(label="About",command=self.AboutMsg)
        helpmenu.add_separator()
        menubar.add_cascade(label="Help",menu=helpmenu)
        self.config(menu=menubar)
        #--------------------|Images|----------------------------------
        current_directory = os.path.dirname(os.path.abspath(__file__))
        image_folder = os.path.join(current_directory, "img")
        box_image_path = os.path.join(image_folder, "box.png")
        trade_image_path = os.path.join(image_folder, "trade.png")
        receiver_image_path = os.path.join(image_folder, "receiver.png")
        withraw_image_path = os.path.join(image_folder, "cash-withdrawal.png")
        send_img = ctk.CTkImage(Image.open(box_image_path), size=(25, 25))
        deposit_img = ctk.CTkImage(Image.open(trade_image_path), size=(25, 25))
        receive_img = ctk.CTkImage(Image.open(receiver_image_path), size=(25, 25))
        withraw_img = ctk.CTkImage(Image.open(withraw_image_path), size=(25, 25))
        #--------------------------|Frames|--------------------------------------
        MainFrame = ctk.CTkFrame(master=self,width=735,height=505,fg_color="white")
        MainFrame.pack()
        MainFrame.place(x=160,y=4)
        BalanceFrame =  ctk.CTkFrame(master=MainFrame,width=350,height=120,fg_color="#051B2B")
        BalanceFrame.pack()
        BalanceFrame.place(x=370,y=14)
        img_path = os.path.join(os.path.dirname(__file__), "img/logo.png")
        img = ctk.CTkImage(light_image=Image.open(img_path),size=(120,120))
        #--------------------------|Labels|---------------------------------------
        qrcode_wallet_address_path = os.path.join(os.path.dirname(__file__), "gorgumu/account/61f2c041d4e9e0f558ae5ee8d6adb62e79bacce45be6d0b6d0bbe8947f5f0dd9.png")
        qrcode_img = ctk.CTkImage(light_image=Image.open(qrcode_wallet_address_path), size=(120, 120))
        qrcode_img_label = ctk.CTkLabel(MainFrame,image=qrcode_img,text="")
        qrcode_img_label.pack()
        qrcode_img_label.place(x=50,y=14)
        img_label = ctk.CTkLabel(BalanceFrame,image=img,text="")
        img_label.pack()
        img_label.place(x=200,y=5)
        totalbalancelabel = ctk.CTkLabel(master=BalanceFrame,text="Balance:",text_color="white")
        totalbalancelabel.pack()
        totalbalancelabel.place(x=40,y=20)
        total_title = ctk.CTkLabel(master=MainFrame,text="Total :",text_color="#051B2B",font=("Arial",14))
        total_title.pack()
        total_title.place(x=250,y=24)
        rubel_balance = self.usd_balance()
        total_value = ctk.CTkLabel(master=MainFrame,text=f"{rubel_balance} $",text_color="#051B2B",font=("Arial",16,"bold"))
        total_value.pack()
        total_value.place(x=250,y=64)
        historylabel = ctk.CTkLabel(master=MainFrame,text="History",text_color="#051B2B",font=("Arial",14))
        historylabel.pack()
        historylabel.place(x=320,y=200)
        countdown = ctk.CTkLabel(master=MainFrame,font=("Arial", 12,"bold"),text="Expire Date:")
        countdown.pack()
        countdown.place(x=30,y=470)
        #----------------------------------|EXPIRE_DATE_WIDGETS|------------------------------------
        m = Myaddress()
        expire_date_query = SimpleStatement(f"SELECT expire_date FROM gorgumu.wallet WHERE address = '{m}';")
        rowx = session.execute(expire_date_query)
        for exp in rowx:
            my_expire_date = exp.expire_date
        countdownvalue = ctk.CTkLabel(master=MainFrame,font=("Arial", 12,"bold"),text=my_expire_date,text_color="coral")
        countdownvalue.pack()
        countdownvalue.place(x=110,y=471)
        date_format = "%Y-%m-%d"
        exp_to_str = str(my_expire_date)
        current_day_to_str = str(datetime.date.today())
        a = datetime.datetime.strptime(exp_to_str, date_format)
        b = datetime.datetime.strptime(current_day_to_str, date_format)
        days_remains = a-b 
        countdownlabel = ctk.CTkLabel(master=MainFrame,font=("Arial", 12,"bold"),text=f"{days_remains.days} Remains")
        countdownlabel.pack()
        countdownlabel.place(x=400,y=472)
        #---------------------------------|BALANCE WIDGETS|-------------------------------------------
        GRM = self.grm_balance()
        tbalance = ctk.CTkLabel(master=BalanceFrame,text=GRM,text_color="white",font=("Arial",26,"bold"))
        tbalance.pack()
        tbalance.place(x=40,y=45)
        #---------------------------------------|Buttons|---------------------------------------------------
        sendtagbutton = ctk.CTkButton(master=MainFrame,text="Send",image=send_img,fg_color="white",corner_radius=5,height=50,font=('Arial', 13),text_color="#051B2B",hover="white",command=self.call_send_window)
        sendtagbutton.pack()
        sendtagbutton.place(x=100,y=150)
        receivetagbutton = ctk.CTkButton(master=MainFrame,text="Receive",image=receive_img,fg_color="white",corner_radius=5,height=50,font=('Arial', 13),text_color="#051B2B",hover="white",command=self.call_receive_window)
        receivetagbutton.pack()
        receivetagbutton.place(x=300,y=150)
        loantagbutton = ctk.CTkButton(master=MainFrame,text="Loan",image=withraw_img,fg_color="white",corner_radius=5,height=50,font=('Arial', 13),text_color="#051B2B",hover="white",command=self.call_loan_window)
        loantagbutton.pack()
        loantagbutton.place(x=500,y=150)
        #-----------------|Treeview|----------------------------------
        history = ttk.Treeview(MainFrame,show="headings")
        history['columns'] = ("Sender","Receipient","Amount","Timesamp")
        history.column("Sender",anchor=W,width=220)
        history.column("Receipient",anchor=W,width=220)
        history.column("Amount",anchor=CENTER,width=80)
        history.column("Timesamp",anchor=W,width=190)
        history.heading("Sender",text="Sender")
        history.heading("Receipient",text="Receipient")
        history.heading("Amount",text="Amount")
        history.heading("Timesamp",text="Timesamp")
        history.pack()
        history.place(x=10,y=240)
        #----------------------|Display My Whole Transactions|------------------------------
        maddress = Myaddress()
        cql = f"SELECT * FROM gorgumu.transaction WHERE sender = '{maddress}' ALLOW FILTERING;"
        rows = session.execute(cql)
        my_transaction_history = []
        for row in rows:
            my_transaction_history.append({'sender':row.sender,"receiver":row.receiver,"amount":row.amount,"datetime":row.datetime})
        if my_transaction_history:
            for i,transaction in enumerate(my_transaction_history):
                history.insert(parent="", index="end", iid=i, text="",values=(transaction['sender'], transaction['receiver'], transaction['amount'], transaction['datetime']))
        else:
            history.insert(parent="", index="end", iid=1, text="",values=("","","",""))
        #-------------------------|Progressbar|----------------------------------
        progressbar = ctk.CTkProgressBar(master=MainFrame,progress_color="#051B2B")
        progressbar["max"] = 120
        progress_value = 120 - days_remains.days
        progressbar.set(progress_value)
        progressbar.pack()
        progressbar.place(x=500,y=482)
        return MainFrame
    #Sideframe
    def sidebarframe(self):
        #-------------------------|Images|------------------------
        current_directory = os.path.dirname(os.path.abspath(__file__))
        image_folder = os.path.join(current_directory, "img")
        settings_image_path = os.path.join(image_folder, "settings.png")
        settings_img = ctk.CTkImage(PIL.Image.open(settings_image_path), size=(25, 25))
        logout_image_path = os.path.join(image_folder, "turn-off.png")
        logout_img = ctk.CTkImage(PIL.Image.open(logout_image_path), size=(20, 20))
        wallet_image_path = os.path.join(image_folder, "wallet.png")
        wallet_img = ctk.CTkImage(PIL.Image.open(wallet_image_path), size=(25, 25))
        #------------------------|Frames|--------------------------
        SideBarFrame = ctk.CTkFrame(master=self,width=150,height=505,fg_color="white")
        SideBarFrame.pack()
        SideBarFrame.place(x=5,y=4)
        #------------------------|Buttons|--------------------------
        PortfolioButton = ctk.CTkButton(master=SideBarFrame,text="Portfolio",fg_color="coral",image=wallet_img,corner_radius=5,height=50,font=('Arial', 15),text_color="#051B2B",hover_color="pink",command=self.mainframe)
        PortfolioButton.pack()
        PortfolioButton.place(x=5,y=10)
        SettingTabButton = ctk.CTkButton(master=SideBarFrame,text="Settings",fg_color="white",image=settings_img,corner_radius=5,height=50,font=('Arial', 15),text_color="#051B2B",hover_color="white",command=self.settings_frame)
        SettingTabButton.pack()
        SettingTabButton.place(x=5,y=80)
        logouttabButton = ctk.CTkButton(master=SideBarFrame,text="Logout",image=logout_img,fg_color="white",corner_radius=5,height=50,font=('Arial', 15),text_color="gray",hover="white",command=self.signout)
        logouttabButton.pack()
        logouttabButton.place(x=5,y=400)
    #settings frame
    def settings_frame(self):
        #-------------------------|Frames|-----------------------------
        settings_window = ctk.CTkFrame(master=self,width=735,height=490,fg_color="white")
        settings_window.pack()
        settings_window.place(x=160,y=4)
        notice2 = ctk.CTkFrame(master=settings_window,width=590,height=70,fg_color="#051B2B")
        notice2.pack()
        notice2.place(x=50,y=280)
        listframes = [self.baseframe,settings_window]
        listframes[0].forget()
        #-------------------------|Labels|--------------------------------
        current_directory = os.path.dirname(os.path.abspath(__file__))
        image_folder = os.path.join(current_directory, "img")
        security_img_path = os.path.join(image_folder, "guard.png")
        security_img = ctk.CTkImage(Image.open(security_img_path), size=(20, 20))
        security_img_label = ctk.CTkLabel(settings_window,image=security_img,text="")
        security_img_label.pack()
        security_img_label.place(x=25,y=30)
        card_img_path = os.path.join(image_folder, "credit-card.png")
        card_img = ctk.CTkImage(Image.open(card_img_path), size=(20, 20))
        card_img_label = ctk.CTkLabel(settings_window,image=card_img,text="")
        card_img_label.pack()
        card_img_label.place(x=25,y=240)
        securitylabel = ctk.CTkLabel(master=settings_window,text="Login & Security:",font=("cursive", 18))
        securitylabel.pack()
        securitylabel.place(x=50,y=30)
        gatewaypaymentlabel = ctk.CTkLabel(master=settings_window,text="Remove Wallet:",font=("Arial", 18))
        gatewaypaymentlabel.pack()
        gatewaypaymentlabel.place(x=50,y=240)
        text = ctk.CTkLabel(master=notice2,text="If You Want Delete Wallet Just Press 'Delete' Button.",text_color="white",font=("Arial",14))
        text.pack()
        text.place(x=15,y=15)
        #---------------------------------------|Entry|---------------------------------------------------
        global newpassword
        newpassword = ctk.CTkEntry(master=settings_window,width=200,height=30,show="*",placeholder_text="Enter New Password:")
        newpassword.pack()
        newpassword.place(x=100,y=110)
        global repeatpassword
        repeatpassword = ctk.CTkEntry(master=settings_window,width=200,height=30,show="*",placeholder_text="Repeate New Password:")
        repeatpassword.pack()
        repeatpassword.place(x=100,y=150)
        #--------------------------------------|Buttons|--------------------------------------------------
        deletebtn = ctk.CTkButton(master=notice2,text="Delete",fg_color="white",hover_color="coral",text_color="#051B2B",command=self.delete_wallet)
        deletebtn.pack()
        deletebtn.place(x=440,y=30)
        savechanges = ctk.CTkButton(master=settings_window,width=200,height=30,fg_color="#051B2B",text="Save",hover_color="coral",command=self.change_password)
        savechanges.pack()
        savechanges.place(x=100,y=190)
    '''
    Function To Let Client Changes The Wallet Password
    '''
    def change_password(self):
        npasswd = newpassword.get()
        repasswd = repeatpassword.get()
        if npasswd == repasswd:
            Mac_address = None
            Hostname = socket.gethostname()
            Operation_System = platform.system()
            Passwd = str(npasswd) + str(Mac_address) + str(Hostname) + str(Operation_System)
            hash = hashlib.sha256(Passwd.encode()).hexdigest()
            with open("core/gorgumu/account/e7cf3ef4f17c3999a94f2c6f612e8a888e5b1026878e4e19398b23bd38ec221a.ini",'w') as newlongpassword:
                newlongpassword.write(hash)
            file_tmp_path = 'core/gorgumu/account/3f3af1ecebbd1410ab417ec0d27bbfcb5d340e177ae159b59fc8626c2dfd9175.tmp'
            os.remove(file_tmp_path)
            messagebox.showinfo('Succufull','Your Password Has Been Changed Succufully!')
            self.destroy()

    '''
    Function To Delete Wallet Address
    '''
    def delete_wallet(self):
        maddress = Myaddress()
        credentials_folder_path = 'core/gorgumu/account'
        response = messagebox.askyesno("Delete Wallet","Do You Want To Remove Your Wallet Serioucely!?")
        if response:
            #----------------------|Delete Wallet From Database|--------------------------
            delete_statement = session.prepare("DELETE FROM gorgumu.wallet WHERE address = ?")
            session.execute(delete_statement, (maddress,))
        #-----------|Delete Keys And Passwords|--------------
        creds_files = os.listdir(credentials_folder_path)
        for cred in creds_files:
            cred_path = os.path.join(credentials_folder_path, cred)
            os.remove(cred_path)
        self.destroy()
    '''
    Function To Sign out From Wallet With Delete One tmp File
    '''
    def signout(self):
        response = messagebox.askquestion("Logout","Are You Sure To Signout?")
        if response == "yes":
            file_path = "core/gorgumu/account/3f3af1ecebbd1410ab417ec0d27bbfcb5d340e177ae159b59fc8626c2dfd9175.tmp"
        try:
            os.remove(file_path)
            print(f"File '{file_path}' deleted successfully.")
            self.destroy()
        except FileNotFoundError:
            print(f"File '{file_path}' not found.")
    '''
    Function To Show Alert Window Of About The Software['Official WeebSite & Coder']
    '''
    def AboutMsg(self):
        messagebox.showinfo("About","Coded By Salah Salah Eddine\nVisit Own Wbsite:http://gorgumu.com")
    '''
    Function To Create Excel File And Store The GRM Price Every Single Day
    '''
    def update_records(self):
        filename = "Records.xlsx"
        filename_path = "core/gorgumu/"
        file_path = os.path.join(filename_path, filename)
        last_update_file = "last_update.txt"
        last_update_file_path = "core/gorgumu/"
        last_update_path = os.path.join(last_update_file_path, last_update_file)
        current_date = datetime.datetime.now().date()
        if os.path.exists(last_update_path):
            with open(last_update_path, 'r') as f:
                last_update = datetime.datetime.strptime(f.read(), '%Y-%m-%d').date()
        else:
            with open(last_update_path, 'w') as f:
                f.write(str(current_date))
            last_update = None
        if current_date != last_update:
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
            else:
                wb = Workbook()
            ws = wb.active
            if last_update is None:
                ws["A1"] = "DATE"
                ws["B1"] = "PRICE"
            ws.append([str(current_date), Gorgumu_Price])
            wb.save(file_path)
            with open(last_update_path, 'w') as f:
                f.write(str(current_date))
    '''
    Function Check If Loan Is Expired
    '''
    def loan_is_expire(self):
        maddress = Myaddress()
        today = datetime.date.today()
        expire_day = session.prepare(f"SELECT expire_date FROM gorgumu.loans WHERE address = '{maddress}';")
        return expire_day == today
    '''
    Function To Paid Your Loan If Expire Date Is Today
        * If Client Paid,His Iscore Rise Up(This Is Helps Client For Demand Another Future Loan)
        * If Not, His Iscore Less Down (This Is Threats Client His Records)
        * The Iscore Rise & Less As Iscore Points (You Can Back To '_finance.py' File To See)
    '''
    def paid_bill(self):
        maddress = Myaddress()
        client_score = my_iscore()
        #---------------------------------|GET CLIEMT FINANCIAL DATA|-------------------------------
        GET_CLIENT_BALANCE_QUERY = SimpleStatement(f"SELECT * FROM gorgumu.wallet WHERE address = '{maddress}';")
        rows1 = session.execute(GET_CLIENT_BALANCE_QUERY)
        for b in rows1:
            CLIENT_BALANCE = b.balance
        GET_LOAN_BOX_AMOUNT_QUERY = SimpleStatement(f"SELECT * FROM gorgumu.loanbox WHERE total = 'MAX';")
        rows2 = session.execute(GET_LOAN_BOX_AMOUNT_QUERY)
        for ls in rows2:
            LOANBOX_AMOUNT = ls.loansupply
        GET_CLIENT_DEBT_QUERY = SimpleStatement(f"SELECT * FROM gorgumu.loans WHERE address = '{maddress}';")
        rows3 = session.execute(GET_CLIENT_DEBT_QUERY)
        for dbt in rows3:
            CLIENT_DEBT = dbt.debt
            if self.loan_is_expire and CLIENT_DEBT is not None:
                #------------------------------|UPDATE Client Iscore|---------------------------------
                NEW_CLIENT_ISCORE = Decimal(client_score) + 10
                UPDATED_CLIENT_ISCORE = session.prepare(f"UPDATE gorgumu.wallet SET iscore = {NEW_CLIENT_ISCORE} WHERE address = '{maddress}'")
                session.execute(UPDATED_CLIENT_ISCORE)
                #------------------------------|MINUS Client Balance From Loan|-------------------------
                NEW_CLIENT_BALANCE_AFTER_LOAN_EXPIRED = Decimal(CLIENT_BALANCE) - Decimal(CLIENT_DEBT)
                CLIENT_UPDATE_BALANCE = session.prepare(f"UPDATE gorgumu.wallet SET balance = {NEW_CLIENT_BALANCE_AFTER_LOAN_EXPIRED} WHERE address = '{maddress}'")
                session.execute(CLIENT_UPDATE_BALANCE)
                #------------------------------|UPDATE MONEYSUPPLY:ADD DEBT TO MONEYSUPPLY|---------------
                NEW_MONEY_SUPPLY_AMOUNT = Decimal(LOANBOX_AMOUNT) + Decimal(CLIENT_DEBT)
                UPDATED_MONEY_SUPPLY = session.prepare(f"UPDATE gorgumu.loanbox SET loansupply = {NEW_MONEY_SUPPLY_AMOUNT} WHERE total = 'MAX'")
                session.execute(UPDATED_MONEY_SUPPLY)
                #------------------------------|UPDATE CLIENT DEBT TO NULL|--------------------------------
                UPDATED_CLIENT_DEBT = session.prepare(f"UPDATE gorgumu.loans SET debt = 0 WHERE address = '{maddress}'")
                session.execute(UPDATED_CLIENT_DEBT)
                messagebox.showinfo("Expired","Your Loan Date Expire Is Today,We Get Back The Amount Succufuly")
            elif CLIENT_BALANCE is None or CLIENT_BALANCE < CLIENT_DEBT and client_score is not None:
                #------------------------------|UPDATE CLIENT ISCORE|---------------------------------------
                NEW_CLIENT_ISCORE = Decimal(client_score) - 50
                UPDATED_CLIENT_ISCORE = session.prepare(f"UPDATE gorgumu.wallet SET iscore = {NEW_CLIENT_ISCORE} WHERE address = '{maddress}'")
                session.execute(UPDATED_CLIENT_ISCORE)
                #-------------------|Start Encryption|-----------
                Encrypt()
            else:
                pass
    '''
    Function Check The Client Wallet Expire Date Is Today
    '''
    def balance_is_expire(self):
        maddress = Myaddress()
        today = datetime.date.today()
        expire_day = session.prepare(f"SELECT expire_date FROM gorgumu.wallet WHERE address = '{maddress}';")
        return expire_day == today
    '''
    Function Update The Client Balance To Zero If The Expire Date Is Today
    '''
    def null_balance(self):
        maddress = Myaddress()
        if self.balance_is_expire():
            update_query = session.prepare(f"UPDATE gorgumu.wallet SET balance = 0 WHERE address = '{maddress}';")
            session.execute(update_query)
            messagebox.showinfo("Balance","Your Balance Null Because Your Expire Date Is Today!")
            print("Your Balance Is Expired")
        else:
            print("The Expire Date Not Comming Yet!")
    #----------------------|Call Windows For Display|---------------------------------
    '''
    Function To Make Event When Client Click Send Tag Button Display The Sender Window
    '''
    def call_send_window(self):
        S = SENDTRXWIN()
        return S
    '''
    Function To Make Event When Client Click Receive Tag Button Display The Receive Window
    '''
    def call_receive_window(self):
        R = RECEIVEWIN()
        return R
    '''
    Function To Make Event When Client Click Loan Tag Button Display The Loan Window
    '''
    def call_loan_window(self):
        L = LOANWIN()
        return L
    '''
    Function Make Event When Client Click On Transaction Tag Button Display Ledger Window
    '''
    def call_ledger_window(self):
        L = LEDGER()
        return L
    '''
    Function Make Event When Client Click On Perfomence Tag Button Display Perfomence Window
    '''
    def call_perfomence_window(self):
        P = PERFOMENCE()
        return P
    '''
    Function Make Event When Client Click On Deposit Tag Button Display Deposite Window
    '''
    def call_deposite_window(self):
        D = DEPOSITE()
        return D
#SENDER TRANSACTION
class SENDTRXWIN(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Send Transaction")
        self.window_width = 500
        self.window_height = 500
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.x = int((self.screen_width/2) - (self.window_width/2))
        self.y = int((self.screen_height/2) - (self.window_height/2))
        self.geometry(f"{self.window_width}x{self.window_height}+{self.x}+{self.y}")
        self.iconbitmap("core/icon/gorgumu_icon.ico")
        self.resizable(False, False)
        self.baseframe = self.mainframe()
        self.mainloop()
    #Widget
    def mainframe(self):
        mainframe = ctk.CTkFrame(master=self,width=500,height=600,fg_color="white")
        mainframe.pack()
        #-----|Label|-------
        addresslabel = ctk.CTkLabel(master=mainframe,text="Wallet Address:",font=("Arial",16,"bold"),text_color="#051B2B")
        addresslabel.pack()
        addresslabel.place(x=20,y=55)
        amountlabel = ctk.CTkLabel(master=mainframe,text="Amount:",font=("Arial",16,"bold"),text_color="#051B2B")
        amountlabel.pack()
        amountlabel.place(x=20,y=145)
        #-----|Entry|------
        global address_to_send
        address_to_send = ctk.CTkEntry(master=mainframe,width=400,height=50,placeholder_text="Enter  Address:")
        address_to_send.pack()
        address_to_send.place(x=30,y=90)
        global amount_to_send
        amount_to_send =  ctk.CTkEntry(master=mainframe,width=400,height=50,placeholder_text="Enter Amount:")
        amount_to_send.pack()
        amount_to_send.place(x=30,y=180)
        #---------------------------------------------------|Combox|----------------------------------------
        options = ["GRM","ZYN"]
        global combobox
        combobox = ctk.CTkComboBox(master=self, values=options,width=400,height=50,dropdown_fg_color="white")
        combobox.pack()
        combobox.place(x=30,y=270)
        #---------------------------------------------------|Buttons|-----------------------------------
        maxamountbtn = ctk.CTkButton(master=mainframe,text="MAX",bg_color="transparent",fg_color="#051B2B",width=50,height=20,text_color="white",hover_color="coral",command=self.maxamount)
        maxamountbtn.pack()
        maxamountbtn.place(x=370,y=195)
        exitbutton = ctk.CTkButton(master=mainframe,width=400,height=50,fg_color="white",text="Exit",text_color="gray",hover_color="white",command=self.destroy)
        exitbutton.pack()
        exitbutton.place(x=30,y=410)
        sendbutton =  ctk.CTkButton(master=mainframe,width=400,height=50,fg_color="coral",text="Send",hover_color="#051B2B",command=self.send_method)
        sendbutton.pack()
        sendbutton.place(x=30,y=350)
    '''
    Function To Get The Whole Balance That Client Has From His Wallet.
    The Client Send This Amount After Clicking 'MAX' And Submit By 'SEND' Button
    '''
    def maxamount(self):
        maddress = Myaddress()
        query = SimpleStatement(f"SELECT balance FROM gorgumu.wallet WHERE address = '{maddress}' ;")
        rows = session.execute(query)
        for row in rows:
            balance = row.balance
        amount_to_send.insert(0,balance)
    '''
    Function Submit The Transaction After Filled The Correct Address & Amount
    **Neccessary Available Internet Connection
    Notice: Before Sent Transaction You Have To Know That One GRM Unit Equal 1024 ZYN
    '''
    def send_method(self):
        connect = is_online()
        try:
            if connect:
                maddress = Myaddress()
                exchange_rate = {"GRM":1,"ZYN":1024}
                currency_selected = combobox.get()
                receipent = address_to_send.get()
                amountt = float(amount_to_send.get())
                if receipent == maddress:
                    return messagebox.showwarning("Failed","The System Not Allow You For Send Amount To Your Address")
                #---|Check If Client Want To Send GRM Amount|----
                elif currency_selected == "GRM":
                    will_send_grm = amountt * exchange_rate["GRM"]
                    #---------|Create Transaction|-----------
                    tx1 = Transaction(receipent,will_send_grm)
                    tx1.calculate_hash()
                    tx1.sign_transaction(tx1.data)
                    tx1.Send(tx1.data,tx1.receiver,tx1.amount)
                    time.sleep(2)
                    self.destroy()
                #---|Check If Client Want To Send ZYN Amount|----
                elif currency_selected == "ZYN":
                    will_send_zyn = amountt / exchange_rate["ZYN"]
                    #----------|Create Transaction|----------
                    tx2 = Transaction(receipent,will_send_zyn)
                    tx2.calculate_hash()
                    tx2.sign_transaction(tx2.data)
                    tx2.Send(tx2.data,tx2.receiver,tx2.amount)
                    time.sleep(2)
                    self.destroy()
            else:
                messagebox.showwarning("Error Connection","Check Your Internet Connection")
        except:
            pass

#Receive Window
class RECEIVEWIN(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Receive")
        self.window_width = 500
        self.window_height = 500
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.x = int((self.screen_width/2) - (self.window_width/2))
        self.y = int((self.screen_height/2) - (self.window_height/2))
        self.geometry(f"{self.window_width}x{self.window_height}+{self.x}+{self.y}")
        self.iconbitmap("core/icon/gorgumu_icon.ico")
        self.baseframe = self.mainframe()
        self.resizable(False, False)
        self.mainloop()
    #Widgets
    def mainframe(self):
        #-------------------------------|Frames|-----------------------------------
        mainframe = ctk.CTkFrame(master=self,width=500,height=600,fg_color="white")
        mainframe.pack()
        warningframe = ctk.CTkFrame(master=self,width=450,height=100,fg_color="#ff9999")
        warningframe.pack()
        warningframe.place(x=20,y=360)
        #------------------------------------------------|Labels|---------------------------------------
        addresslabel = ctk.CTkLabel(master=mainframe,text="Wallet Address:",font=("Arial",16,"bold"),text_color="#051B2B")
        addresslabel.pack()
        addresslabel.place(x=30,y=170)
        privateaddresslabel = ctk.CTkLabel(master=mainframe,text="Private Address:",font=("Arial",16,"bold"),text_color="#051B2B")
        privateaddresslabel.pack()
        privateaddresslabel.place(x=30,y=260)
        alert = f"WARNING:Dont Share Private Key.If Someone Knows This Key Can Send\n Transaction From Your Wallet.You Are The Only Responsable"
        warningalert = ctk.CTkLabel(master=warningframe,text=alert,font=("Arial",12,"bold"),text_color="white")
        warningalert.pack()
        warningalert.place(x=15,y=35)
        #------------------------------------------------|Entry|----------------------------------------
        global address
        address = ctk.CTkEntry(master=mainframe,width=400,height=50,placeholder_text="",corner_radius=5,text_color="#051B2B",font=("Arial",14,'bold'))
        address.pack()
        address.place(x=30,y=200)
        wallet_address = Myaddress()
        address.insert(1,wallet_address)
        privateaddress = ctk.CTkEntry(master=mainframe,width=400,height=50,placeholder_text="",corner_radius=5,text_color="Gray",font=("Arial",14,"bold"))
        privateaddress.pack()
        privateaddress.place(x=30,y=290)
        privateaddress.insert(1,"PRIVATE ADDRESS")
        private_key = Private_key()
        privateaddress.insert(1,private_key)
        privateaddress.configure(state="disabled")
        #-------------------------------------------------|Buttons|-----------------------------------------
        global copyaddressbtn
        copyaddressbtn = ctk.CTkButton(master=mainframe,text="Copy",width=50,height=30,hover_color="coral",command=self.copyaddress,fg_color="#051B2B")
        copyaddressbtn.pack()
        copyaddressbtn.place(x=375,y=210)
    '''
    Function To Client Can Copy His Wallet Address For Give it To Somebody For Receive Balance From Him
    '''
    def copyaddress(self):
        pyperclip.copy(address.get())
        copyaddressbtn.configure(text="Copied")
#Loan Window
class LOANWIN(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.window_width = 500
        self.window_height = 500
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.x = int((self.screen_width/2) - (self.window_width/2))
        self.y = int((self.screen_height/2) - (self.window_height/2))
        self.geometry(f"{self.window_width}x{self.window_height}+{self.x}+{self.y}")
        self.iconbitmap("core/icon/gorgumu_icon.ico")
        self.title("Loans")
        self.baseframe = self.mainframe()
        self.resizable(False, False)
        self.mainloop()
    #Widgets
    def mainframe(self):
        mainframe = ctk.CTkFrame(master=self,width=500,height=600,fg_color="white")
        mainframe.pack()
        #--------------------------------------------------|Label|-------------------------------------------
        amountlabel = ctk.CTkLabel(master=mainframe,text="Amount:",font=("Arial",16,"bold"),text_color="#051B2B")
        amountlabel.pack()
        amountlabel.place(x=20,y=12)
        term1 = ctk.CTkLabel(master=mainframe,text="1-At Least 20 Transactions",font=("Arial",13),text_color="#051B2B")
        term1.pack()
        term1.place(x=35,y=110)
        term2 = ctk.CTkLabel(master=mainframe,text="2-Max Day Limit : 60 Days ",font=("Arial",13),text_color="#051B2B")
        term2.pack()
        term2.place(x=35,y=140)
        term3 = ctk.CTkLabel(master=mainframe,text="3-Type Loan:Ninja (Client Data is Asset)",font=("Arial",13),text_color="#051B2B")
        term3.pack()
        term3.place(x=35,y=170)
        term4 = ctk.CTkLabel(master=mainframe,text="4-Fractional Reserve:10% ",font=("Arial",13),text_color="#051B2B")
        term4.pack()
        term4.place(x=35,y=200)
        #----------------------------------------------------|Entry|-----------------------------------------------
        global loanamount
        loanamount = ctk.CTkEntry(master=mainframe,width=400,height=50,font=("Arial",18,"bold"),text_color="gray",placeholder_text="Enter Amount You Want It")
        loanamount.pack()
        loanamount.place(x=30,y=50)
        #-----------------------------------------------------|Buttons|---------------------------------------------
        exitbutton = ctk.CTkButton(master=mainframe,width=400,height=50,fg_color="white",text="Exit",text_color="gray",hover_color="white",command=self.destroy)
        exitbutton.pack()
        exitbutton.place(x=30,y=410)
        genbutton =  ctk.CTkButton(master=mainframe,width=400,height=50,fg_color="coral",text="Agree & Generate",hover_color="#051B2B",command=self.loan_method)
        genbutton.pack()
        genbutton.place(x=30,y=350)
    '''
    Function Give Client a Method To Get Loan Type Ninja With Conditions
    '''
    def loan_method(self):
        maddress = Myaddress()
        loanbox_query = SimpleStatement("SELECT * FROM gorgumu.loanbox;")
        rows = session.execute(loanbox_query)
        for rs in rows:
            loansupply = rs.loansupply
        #Allow Client To Get Just Half Of Total Loansupply
        global halfsupply
        halfsupply = loansupply / 2
        loanamountvalue = Decimal(loanamount.get())
        if loanamountvalue <= halfsupply:
            return calculate_loan(maddress,loanamountvalue)
        else:
            messagebox.showwarning("Warning","Something Wrong! We Dont Have This Supply")
#Ledger
class LEDGER(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.window_width = 330
        self.window_height = 380
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.x = int((self.screen_width/2) - (self.window_width/2))
        self.y = int((self.screen_height/2) - (self.window_height/2))
        self.geometry(f"{self.window_width}x{self.window_height}+{self.x}+{self.y}")
        self.title("Ledger")
        self.iconbitmap("core/icon/gorgumu_icon.ico")
        self.widget = self.widgets()
        self.resizable(False, False)
        self.mainloop()
    #Widgets
    def widgets(self):
        #---------------|Treeview|-----------------
        history = ttk.Treeview(self,show="headings")
        history['columns'] = ("Transaction_Id")
        history.pack()
        history.place(x=5,y=10,width=315, height=360)
        history.column("Transaction_Id",anchor=W,width=150)
        history.heading("Transaction_Id",text="TXID")
        query = f"SELECT * FROM gorgumu.transaction;"
        rows = session.execute(query)
        ledger_transaction = []
        for row in rows:
            ledger_transaction.append({'Transaction_Id':row.hash})
        if ledger_transaction:
            for i,transaction in enumerate(ledger_transaction):
                history.insert(parent="", index="end", iid=i, text="",values=(transaction['Transaction_Id']))
        else:
            history.insert(parent="", index="end", iid=1, text="",values=("-"))
#Deposite
class DEPOSITE(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.window_width = 500
        self.window_height = 500
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.x = int((self.screen_width/2) - (self.window_width/2))
        self.y = int((self.screen_height/2) - (self.window_height/2))
        self.geometry(f"{self.window_width}x{self.window_height}+{self.x}+{self.y}")
        self.iconbitmap("core/icon/gorgumu_icon.ico")
        self.resizable(False, False)
        self.title("Deposit")
        self.baseframe = self.mainframe()
        self.mainloop()
    #Widets
    def mainframe(self):
        #------------------------------|Frames|----------------------------------
        mainframe4 = ctk.CTkFrame(master=self,width=500,height=600,fg_color="white")
        mainframe4.pack()
        #------------------------|Menubar|---------------------------
        menubar = Menu(self)
        menu = Menu(menubar, tearoff=0)
        menu.add_command(label="GET CASH",command=self.get_back_cash)
        menu.add_separator()
        menubar.add_cascade(label="GET",menu=menu)
        self.config(menu=menubar)
        #---------------------------------------------------|Lables|--------------------------------------------------
        addresslabel = ctk.CTkLabel(master=mainframe4,text="Wallet Address:",font=("Arial",16,"bold"),text_color="#051B2B")
        addresslabel.pack()
        addresslabel.place(x=30,y=170)
        amountinputlabel = ctk.CTkLabel(master=mainframe4,text="Amount:",font=("Arial",16,"bold"),text_color="#051B2B")
        amountinputlabel.pack()
        amountinputlabel.place(x=30,y=260)
        #----------------------------------------------------|Entry|------------------------------------------------------
        maddress = Myaddress()
        useraddress = ctk.CTkEntry(master=mainframe4,width=400,height=50,placeholder_text="",corner_radius=5,text_color="#051B2B",font=("Arial",14,'bold'))
        useraddress.pack()
        useraddress.place(x=30,y=200)
        useraddress.insert(1,maddress)
        useraddress.configure(state="disabled")
        global amountdeposit
        amountdeposit = ctk.CTkEntry(master=mainframe4,width=400,height=50,placeholder_text="Enter Amount To Deposit:",corner_radius=5,text_color="#051B2B",font=("Arial",14,'bold'))
        amountdeposit.pack()
        amountdeposit.place(x=30,y=290)
        #----------------------------------------------------|Buttons|------------------------------------------------------
        depositbutton =  ctk.CTkButton(master=mainframe4,width=400,height=50,fg_color="coral",text="Deposit",hover_color="#051B2B",command=self.deposite_method)
        depositbutton.pack()
        depositbutton.place(x=30,y=360)
    '''
    Function To Get Back The Cash That Client Deposite If Available On Reserve
    '''
    def get_back_cash(self):
        maddress = Myaddress()
        #--------------------------|Get Client Balance|-------------------------------
        query = SimpleStatement(f"SELECT * FROM gorgumu.wallet WHERE address = '{maddress}' ;")
        rows = session.execute(query)
        for b in rows:
            my_balance = b.balance
        #---------|Get The Amount That Client Deposite From DEPOSITE Table|-----------------
        query1 = SimpleStatement(f"SELECT amount FROM gorgumu.DEPOSITE WHERE address = '{maddress}' ;")
        rows1 = session.execute(query1)
        for r in rows1:
            MCash = r.amount
        #-------|Check How Much Amount On Fractional Reserve|--------
        query2 = SimpleStatement("SELECT * FROM gorgumu.FRACTIONAL_RESERVE ;")
        rows2 = session.execute(query2)
        for supply in rows2:
            #Whole FRACTIONAL RESERVE Amount
            Total_Reserve = supply.money_supply
        #Check If Client Deposit Amount Not None
        if MCash is not None and MCash < Total_Reserve:
            BACKCASH = Total_Reserve - Decimal(MCash)
            NEWBALANCE = Decimal(my_balance) + Decimal(MCash)
            #-----------------------------------------------|UPDATE CLIENT BALANCE|--------------------------------
            query3 = session.prepare(f"UPDATE gorgumu.wallet SET balance = {NEWBALANCE} WHERE address = '{maddress}'")
            session.execute(query3)
            #-----------------------------------------------|UPDATE FRACTIONAL_RESERVE|----------------------------
            FRACTIONAL_RESERVE_UPDATE_QUERY = session.prepare(f"UPDATE gorgumu.FRACTIONAL_RESERVE SET money_supply = {BACKCASH} WHERE total = 'MAX'")
            session.execute(FRACTIONAL_RESERVE_UPDATE_QUERY) 
            messagebox.showinfo("Succuful","Your Amount That You Deposit Has Been Whitrawed Succufuul.Thank You For Trust Us")
        if MCash == '':
            messagebox.showwarning("Failed","Your Amount Didn't Found Or Dosen't Available Currently!")
    '''
    Function Let Client Deposite His Amount For Store It
    The System Use Fractional Reserve
    '''
    def deposite_method(self):
        maddress = Myaddress()
        query = SimpleStatement(f"SELECT balance FROM gorgumu.wallet WHERE address = '{maddress}' ;")
        rows = session.execute(query)
        for row in rows:
            Obalance = row.balance 
        my_amount_to_deposit = Decimal(amountdeposit.get())
        #-------------|Check To Confirm My Amount Wanna Deposit it Dosen't Big Then My Balance|---------
        if my_amount_to_deposit > Decimal(Obalance):
            amountdeposit.configure(border_color="red")
            amountdeposit.place(x=30,y=260)
            warningdeposit = ctk.CTkLabel(master=self,text="You Dont Have This Balance!",text_color="red")
            warningdeposit.pack()
            warningdeposit.place(x=30,y=320)
            messagebox.showwarning("Warning","You Dont Have This Big Balance! ")
        #-------------|Return False When Amount Is Null|---------------------------
        elif my_amount_to_deposit == 0:
            amountdeposit.configure(border_color="red")
            messagebox.showerror("Error","You Cant Enter Null Amount.Try Again")
        #--|If True Deposit , Insert Wallet Address & Amount To Deposit Table--|
        elif my_amount_to_deposit <= Obalance:
            #---------------------|UPDATE Client Balance|-----------------------
            NEW_CLIENT_BALANCE = Decimal(Obalance) - Decimal(my_amount_to_deposit)
            QUERY = SimpleStatement(f"UPDATE gorgumu.wallet SET balance = {NEW_CLIENT_BALANCE} WHERE address = '{maddress}'")
            session.execute(QUERY)
            #-------------|CALCULATE 10% FROM DEPOSITE Amount FOR Fractional Reserve Amount|------
            FRACTIONAL_RESERVE_RATE = 10
            CALCULATE_PERCENTAGE = FRACTIONAL_RESERVE_RATE / 100
            FRACTIONAL_RESERVE = Decimal(my_amount_to_deposit) * Decimal(CALCULATE_PERCENTAGE)
            #------------------|CALCULATE LOAN AMOUNT|-----------------
            LOAN_VALUE = Decimal(my_amount_to_deposit) - FRACTIONAL_RESERVE
            FRACTIONAL_RESERVE_RESULT = Decimal(my_amount_to_deposit) - LOAN_VALUE
            #------------------|SELECT LOANBOX & FRACTIONAL_RESERVE|----------------
            get_loanbox_amount = SimpleStatement("SELECT * FROM gorgumu.Loanbox WHERE total = 'MAX';")
            row = session.execute(get_loanbox_amount)
            for x in row:
                loanbox_amout = x.loansupply
            get_fractional_reserve_amount = SimpleStatement("SELECT * FROM gorgumu.Fractional_Reserve WHERE total = 'MAX';")
            rows = session.execute(get_fractional_reserve_amount)
            for ys in rows:
                fractional_reserve_x = ys.money_supply
            #-----------------------------|INSERT Deposit Amount To "DEPOSITE" TABLE|--------------------------
            INSERT_DEPOSITE = session.prepare(f"INSERT INTO gorgumu.DEPOSITE (address,amount) VALUES ('{maddress}',{my_amount_to_deposit})")
            session.execute(INSERT_DEPOSITE)
            #--------|CALCULATE 10% FROM DEPOSITE Amount FOR Fractional Reserve Amount|-------
            FRACTIONAL_RESERVE_RATE = 10
            CALCULATE_PERCENTAGE = FRACTIONAL_RESERVE_RATE / 100
            FRACTIONAL_RESERVE = Decimal(my_amount_to_deposit) * Decimal(CALCULATE_PERCENTAGE)
            #|-------|CALCULATE LOAN AMOUNT|-------------------
            LOAN_VALUE = my_amount_to_deposit - FRACTIONAL_RESERVE
            UPDATE_LOAN_VALUE = loanbox_amout + Decimal(LOAN_VALUE)
            #|-----------------------------------|INSERT FRACTIONAL RESERVE & LOAN INTO TABLES|-----------------------
            INSERT_LOAN = session.prepare(f"UPDATE gorgumu.loanbox SET loansupply = {UPDATE_LOAN_VALUE} WHERE total = 'MAX';")
            session.execute(INSERT_LOAN)
            FRACTIONAL_RESERVE_RESULT = my_amount_to_deposit - LOAN_VALUE
            UPDATED_FRACTIONAL_RESERVE = fractional_reserve_x + Decimal(FRACTIONAL_RESERVE_RESULT)
            INSERT_FRACTIONAL_RESERVE = session.prepare(f"UPDATE gorgumu.Fractional_Reserve SET money_supply = {UPDATED_FRACTIONAL_RESERVE} WHERE total = 'MAX';")
            session.execute(INSERT_FRACTIONAL_RESERVE)
            time.sleep(1)
            self.destroy()
            messagebox.showinfo("Deposit","Deposit Succufull Thank You!")
        
#Perfomence
class PERFOMENCE(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.window_width = 350
        self.window_height = 200
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.x = int((self.screen_width/2) - (self.window_width/2))
        self.y = int((self.screen_height/2) - (self.window_height/2))
        self.geometry(f"{self.window_width}x{self.window_height}+{self.x}+{self.y}")
        self.title("Perfomence CPU / RAM")
        self.iconbitmap("core/icon/gorgumu_icon.ico")
        self.resizable(False, False)
        self.baseframe = self.mainframe()
        self.mainloop()
    #Widgets
    def mainframe(self):
        #--------------------------------|Label|--------------------------------------
        cputitle = ctk.CTkLabel(master=self,text="CPU Usage:",font=("Arial",14,"bold"))
        cputitle.pack()
        cputitle.place(x=130,y=20)
        cpupercent = ctk.CTkLabel(master=self,text=f"{psutil.cpu_percent()} %",text_color="red")
        cpupercent.pack()
        cpupercent.place(x=150,y=50)
        ramtitle = ctk.CTkLabel(master=self,text="RAM Usage:",font=("Arial",14,"bold"))
        ramtitle.pack()
        ramtitle.place(x=130,y=85)
        rampercent = ctk.CTkLabel(master=self,text=f"{psutil.virtual_memory().percent} %",text_color="red")
        rampercent.pack()
        rampercent.place(x=150,y=120)
#Signup
class Signin(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.window_width = 750
        self.window_height = 500
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.x = int((self.screen_width/2) - (self.window_width/2))
        self.y = int((self.screen_height/2) - (self.window_height/2))
        self.geometry(f"{self.window_width}x{self.window_height}+{self.x}+{self.y}")
        self.iconbitmap("core/icon/gorgumu_icon.ico")
        self.title("Gorgumu v 1.0")
        self.resizable(False, False)
        self.leftframe = self.pictureframe()
        self.mainloop()
    #Widget
    def pictureframe(self):
        pictureframe = ctk.CTkFrame(master=self,width=350,height=500,corner_radius=None)
        pictureframe.pack()
        pictureframe.place(x=1,y=1)
        global walletframe
        walletframe = ctk.CTkFrame(master=self,width=398,height=500,fg_color="white")
        walletframe.pack()
        walletframe.place(x=350,y=1)
        #------------|Label|------------
        image_path = "core/img/cover.png"
        image = Image.open(image_path)
        photo_image = ImageTk.PhotoImage(image)
        label = ctk.CTkLabel(master=pictureframe, image=photo_image)
        label.configure(text="")
        label.pack()
        welcomelabel = ctk.CTkLabel(master=walletframe,text="Welcome To Gorgumu",text_color="black",font=("Arial",14))
        welcomelabel.pack()
        welcomelabel.place(x=150,y=10)
        rights = ctk.CTkLabel(master=walletframe,text="All Rights Reserved",text_color="Gray")
        this_year = datetime.datetime.now()
        year = ctk.CTkLabel(master=walletframe,text=this_year.year,text_color="Gray")
        rights.pack()
        year.pack()
        rights.place(x=150,y=400)
        year.place(x=180,y=420)
        #-------------------------------------------------|Entry|-----------------------------------------
        global pswds
        pswds = ctk.CTkEntry(master=walletframe,placeholder_text="Enter Your Password:",fg_color="white",border_color="white",corner_radius=20,height=50,width=300,text_color="#051B2B",show="*")
        pswds.pack()
        pswds.place(x=50,y=190)
        #-------------------------------------------------|Buttons|----------------------------------------
        unlock = ctk.CTkButton(master=walletframe,text="Unlock",fg_color="#051B2B",corner_radius=20,width=300,height=50,hover_color="coral",command=self.login_method,font=("Arial",12,"bold"))
        unlock.pack()
        unlock.place(x=50,y=250)
        signup = ctk.CTkButton(master=walletframe,text="Register",fg_color="coral",corner_radius=20,width=300,height=50,hover_color="#051B2B",command=self.signup_frame,font=("Arial",12,"bold"))
        signup.pack()
        signup.place(x=50,y=320)
    '''
    Function To Display The Signup Frame For Register Or Create a New Wallet
    '''
    def signup_frame(self):
        walletframe = ctk.CTkFrame(master=self,width=398,height=500,fg_color="white")
        walletframe.pack()
        walletframe.place(x=350,y=1)
        #-------------------------------------------|Lable|------------------------------------------
        welcomelabel = ctk.CTkLabel(master=walletframe,text="Create A New Wallet",text_color="#051B2B")
        welcomelabel.pack()
        welcomelabel.place(x=150,y=10)
        regex = ctk.CTkLabel(master=walletframe,text="Password Character Must Be:",text_color="#051B2B")
        regex.pack()
        regex.place(x=50,y=190)
        regex_uppercase = ctk.CTkLabel(master=walletframe,text="One Uppercase Minimum",text_color="coral")
        regex_uppercase.pack()
        regex_uppercase.place(x=70,y=260)
        regex_symbol = ctk.CTkLabel(master=walletframe,text="One Symbol Minimum",text_color="coral")
        regex_symbol.pack()
        regex_symbol.place(x=70,y=220)
        regex_numbers = ctk.CTkLabel(master=walletframe,text="Three Numbers Minimum",text_color="coral")
        regex_numbers.pack()
        regex_numbers.place(x=70,y=240)
        #--------------------------------------------|Entry|----------------------------------------------
        global password
        password = ctk.CTkEntry(master=walletframe,placeholder_text="Enter Your Password:",fg_color="white",border_color="white",corner_radius=0,height=50,width=300,text_color="#051B2B",show="*")
        password.pack()
        password.place(x=50,y=100)
        #-------------------------------------|Call Wallet Class From Client.py|--------------------------
        global signup
        signup = Wallet()
        global repassword
        repassword = ctk.CTkEntry(master=walletframe,placeholder_text="Retype Your Password:",fg_color="white",border_color="white",corner_radius=0,height=50,width=300,text_color="#051B2B",show="*")
        repassword.pack()
        repassword.place(x=50,y=140)
        #-----------------------------------------------|Buttons|-------------------------------------------
        register = ctk.CTkButton(master=walletframe,text="Register",fg_color="#051B2B",corner_radius=20,width=300,height=50,hover_color="coral",command=self.signup_method,font=("Arial",12,"bold"))
        register.pack()
        register.place(x=50,y=300)
        back = ctk.CTkButton(master=walletframe,text="Back",text_color="gray",fg_color="white",corner_radius=20,width=300,height=50,hover_color="white",command=self.pictureframe,font=("Arial",12,"bold"))
        back.pack()
        back.place(x=50,y=355)
    '''
    Function To Create a New Wallet
    '''
    def signup_method(self):
        x = password.get()
        if not password.get() and not repassword.get():
            return False
        if password.get() == repassword.get():
            signup.save_keys()
            signup.encrypt_password(x)
            signup.register_wallet()
            time.sleep(5)
            self.destroy()
            signup.myqrcode()
            Main()
    '''
    Function To Log The Client In Wallet After Type His Own Password
    '''
    def login_method(self):
        psl = pswds.get()
        with open("core/gorgumu/account/e7cf3ef4f17c3999a94f2c6f612e8a888e5b1026878e4e19398b23bd38ec221a.ini",'r') as psw:
            readpasswd = psw.read()
        Mac_address = None
        Hostname = socket.gethostname()
        Operation_System = platform.system()
        Passwd = str(psl) + str(Mac_address) + str(Hostname) + str(Operation_System)
        hash = hashlib.sha256(Passwd.encode()).hexdigest()
        if readpasswd == hash:
            print("Login Succufully")
            #------------------------------------|Save Session File Again To Keep App Open When it Found|-----------------------
            today = datetime.datetime.now()
            with open("core/gorgumu/account/3f3af1ecebbd1410ab417ec0d27bbfcb5d340e177ae159b59fc8626c2dfd9175.tmp","w") as Session:
                Session.write(str(today))
            self.destroy()
            Main()
            
        else:
            print("Login Failed")
            pswds.configure(border_color="red")
            pswds.place(x=50,y=160)
            wrong_passwd_label = ctk.CTkLabel(master=walletframe,text="Wrong Password Try Again!",text_color="red")
            wrong_passwd_label.place(x=100,y=218)
        